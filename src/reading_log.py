"""
reading_log.py — the running .xlsx reading log (Zotero-lite).

The spreadsheet is the SINGLE SOURCE OF TRUTH for "already seen": we read its
Title/DOI to know what not to re-feature. New rows are appended; existing rows
(including the user-edited "Read" column) are never touched.

Columns:
  Date Added | Title | Authors | Venue | Year | DOI/URL | Section |
  Why It's Relevant | Paper of the Day | Full Text Used | Read

Only "Read" is user-editable. Everything else is written once and left alone.
"""
from __future__ import annotations
import datetime as dt
import logging
import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

log = logging.getLogger("readinglog")

COLUMNS = ["Date Added", "Title", "Authors", "Venue", "Year", "DOI/URL",
           "Section", "Why It's Relevant", "Paper of the Day",
           "Full Text Used", "Read"]

HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
POTD_FILL = PatternFill("solid", fgColor="FFF3D6")


def _norm_title(t):
    return re.sub(r"[^a-z0-9]+", "", (t or "").lower())


def _norm_doi(doi):
    doi = (doi or "").lower().strip()
    # strip any doi.org URL prefix so bare-DOI and URL forms match
    doi = re.sub(r"^https?://(dx\.)?doi\.org/", "", doi)
    return doi


def _seen_key(title, doi):
    doi = _norm_doi(doi)
    # only treat it as a DOI key if it actually looks like a DOI
    if doi.startswith("10."):
        return f"doi:{doi}"
    return f"title:{_norm_title(title)}"


def load_seen(path: str) -> set:
    """Return the set of already-seen keys from an existing log (empty if none)."""
    if not os.path.exists(path):
        return set()
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(min_row=2, values_only=True)
        seen = set()
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        ti = header.index("Title") if "Title" in header else 1
        di = header.index("DOI/URL") if "DOI/URL" in header else 5
        for r in rows:
            if not r:
                continue
            title = r[ti] if len(r) > ti else ""
            doi = r[di] if len(r) > di else ""
            if title:
                seen.add(_seen_key(str(title), str(doi or "")))
        wb.close()
        return seen
    except Exception as e:
        log.warning("could not read existing log (%s); treating as empty", e)
        return set()


def _year_of(date_str):
    return (date_str or "")[:4]


def append_rows(path: str, new_papers: list[dict]) -> int:
    """Append new_papers to the log (creating it if absent), preserving all
    existing rows/edits. Each paper dict needs: title, authors(list), venue,
    date, url/doi, section, note, is_potd(bool), full_text_used(bool).
    Returns number of rows appended."""
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reading Log"
        ws.append(COLUMNS)
        for ci, _ in enumerate(COLUMNS, start=1):
            c = ws.cell(row=1, column=ci)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        _set_widths(ws)

    today = dt.date.today().isoformat()
    added = 0
    for p in new_papers:
        authors = p.get("authors") or []
        author_str = (authors[0] + " et al." if len(authors) > 1
                      else (authors[0] if authors else ""))
        row = [
            today,
            p.get("title", ""),
            author_str,
            p.get("venue", ""),
            _year_of(p.get("date", "")),
            p.get("url") or p.get("doi", ""),
            p.get("section", ""),
            p.get("note", ""),
            "Yes" if p.get("is_potd") else "",
            "Yes" if p.get("full_text_used") else "No",
            "No",  # Read — user edits this later
        ]
        ws.append(row)
        if p.get("is_potd"):
            for ci in range(1, len(COLUMNS) + 1):
                ws.cell(row=ws.max_row, column=ci).fill = POTD_FILL
        added += 1

    wb.save(path)
    return added


def _set_widths(ws):
    widths = [12, 46, 20, 20, 6, 34, 32, 50, 10, 12, 6]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
